import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import re
import unicodedata
from datetime import date, datetime

from modulos.campos_opcionales import obtener_campos_activos
from gui.ui_helpers import bloquear_columnas, configurar_ventana


class ImportarExcelWindow:
    """
    Importación masiva de productos desde Excel.

    Validaciones ANTES de importar:
      1. Columnas mínimas requeridas (nombre o codigo + precio_unitario)
      2. Duplicados DENTRO del Excel: mismo codigo O mismo nombre en más de una fila
      3. Filas con datos inválidos (precio negativo, cantidad no numérica, etc.)

    En la BD la detección de existente usa:
      - Primero busca por 'codigo' (si se proporcionó)
      - Si no, busca por 'nombre'
    """

    COLUMNAS_BASE = ['codigo', 'nombre', 'descripcion', 'cantidad',
                     'precio_compra', 'precio_unitario', 'stock_minimo', 'stock_maximo',
                     'unidad_medida', 'proveedor', 'categoria']

    ALIASES_COLUMNAS_BASE = {
        'codigo': {'codigo', 'cod', 'sku', 'codigo_sku', 'codigo_producto'},
        'nombre': {'nombre', 'producto', 'nombre_producto'},
        'descripcion': {'descripcion', 'detalle', 'detalle_producto'},
        'cantidad': {'cantidad', 'stock', 'existencia', 'existencias'},
        'precio_compra': {'precio_compra', 'precio_costo', 'costo', 'costo_unitario'},
        'precio_unitario': {'precio_unitario', 'precio', 'precio_venta', 'venta_unitaria', 'pvp'},
        'stock_minimo': {'stock_minimo', 'stock_min', 'minimo', 'stock_bajo'},
        'stock_maximo': {'stock_maximo', 'stock_max', 'maximo', 'tope_stock'},
        'unidad_medida': {'unidad_medida', 'unidad', 'medida', 'unidad_inventario'},
        'proveedor': {'proveedor', 'suplidor', 'supplier'},
        'categoria': {'categoria', 'rubro', 'tipo', 'familia'},
    }

    def __init__(self, master, db, C, usuario):
        self.db      = db
        self.C       = C
        self.usuario = usuario
        self._datos  = []          # filas limpias listas para importar
        self._problemas = []       # lista de (fila, tipo, detalle) con problemas detectados
        self._campos_opcionales = obtener_campos_activos(db)
        self.COLUMNAS = self.COLUMNAS_BASE + [c['key'] for c in self._campos_opcionales]
        self._aliases_columnas = self._construir_aliases_columnas()

        self.win = tk.Toplevel(master)
        self.win.title("📥 Importar Productos desde Excel")
        self.win.configure(bg=C['bg'])
        self.win.grab_set()
        configurar_ventana(self.win, size='main', min_width=1200, min_height=760, start_maximized=True)
        self._build()

    # ── UI ────────────────────────────────────────────────────────────────────

    def _texto_columnas_reconocidas(self):
        columnas = list(self.COLUMNAS_BASE)
        columnas.extend(c['key'] for c in self._campos_opcionales)
        etiquetas = []
        for col in columnas:
            if col == 'codigo':
                etiquetas.append('codigo (opcional)')
            elif col in ('nombre', 'precio_unitario'):
                etiquetas.append(f'{col}*')
            else:
                etiquetas.append(col)
        return ', '.join(etiquetas)

    @staticmethod
    def _normalizar_valor_excel(valor):
        if isinstance(valor, datetime):
            return valor.strftime('%Y-%m-%d')
        if isinstance(valor, date):
            return valor.strftime('%Y-%m-%d')
        return valor

    @staticmethod
    def _normalizar_cabecera(valor):
        texto = unicodedata.normalize('NFKD', str(valor or ''))
        texto = texto.encode('ascii', 'ignore').decode('ascii').lower().strip()
        texto = re.sub(r'[^a-z0-9]+', '_', texto).strip('_')
        return texto

    def _construir_aliases_columnas(self):
        aliases = {
            clave: {self._normalizar_cabecera(alias) for alias in valores}
            for clave, valores in self.ALIASES_COLUMNAS_BASE.items()
        }
        for campo in self._campos_opcionales:
            aliases.setdefault(campo['key'], set()).update({
                self._normalizar_cabecera(campo['key']),
                self._normalizar_cabecera(campo['label']),
                self._normalizar_cabecera(campo.get('col_tabla', '')),
            })
        return aliases

    def _mapear_cabecera(self, valor):
        cabecera = self._normalizar_cabecera(valor)
        if not cabecera:
            return None
        for columna, aliases in self._aliases_columnas.items():
            if cabecera in aliases:
                return columna
        return cabecera if cabecera in self.COLUMNAS else None

    def _ejemplos_plantilla(self):
        ejemplos = [
            {
                'codigo': 'PROD-001', 'nombre': 'Monitor LG 24"',
                'descripcion': 'Full HD IPS 75Hz', 'cantidad': 10,
                'precio_unitario': 250.00, 'stock_minimo': 5, 'stock_maximo': 40,
                'unidad_medida': 'Unidad', 'proveedor': 'LG Electronics',
                'categoria': 'Electrónica',
            },
            {
                'codigo': 'PROD-002', 'nombre': 'Teclado Mecánico',
                'descripcion': 'Switch Blue, retroiluminado', 'cantidad': 15,
                'precio_unitario': 85.00, 'stock_minimo': 3, 'stock_maximo': 30,
                'unidad_medida': 'Unidad', 'proveedor': 'Logitech',
                'categoria': 'Periféricos',
            },
        ]

        for campo in self._campos_opcionales:
            key = campo['key']
            valores = {
                'codigo_barras': ('7501234567890', '7501234567891'),
                'marca': ('LG', 'Logitech'),
                'modelo': ('24MP400', 'K552'),
                'color': ('Negro', 'Negro'),
                'peso': (3.2, 0.9),
                'ubicacion': ('A-01', 'B-03'),
                'numero_serie': ('SN-MON-001', 'SN-TEC-002'),
                'garantia_meses': (12, 6),
                'fecha_vence': ('', ''),
                'impuesto_pct': (15, 15),
            }.get(key, ('', ''))
            ejemplos[0][key], ejemplos[1][key] = valores

        return ejemplos

    def _build(self):
        C = self.C

        ttk.Label(self.win, text="📥 Importar Productos desde Excel",
                  style='Header.TLabel').pack(pady=(14, 2), padx=16, anchor='w')
        ttk.Label(self.win,
                  text=f"Columnas reconocidas: {self._texto_columnas_reconocidas()}",
                  foreground='#64748B', wraplength=900).pack(padx=16, anchor='w')

        # Botones superiores
        frm_top = ttk.Frame(self.win)
        frm_top.pack(fill=tk.X, padx=16, pady=10)

        ttk.Button(frm_top, text="📄 Descargar Plantilla",
                   command=self._descargar_plantilla,
                   style='Neutral.TButton').pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(frm_top, text="📂 Cargar Archivo Excel",
                   command=self._cargar_archivo).pack(side=tk.LEFT)

        self.lbl_archivo = ttk.Label(frm_top, text="Ningún archivo cargado",
                                      foreground='#64748B')
        self.lbl_archivo.pack(side=tk.LEFT, padx=14)

        # Notebook: Previsualización | Problemas detectados
        self.nb = ttk.Notebook(self.win)
        self.nb.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 8))

        self.tab_prev = ttk.Frame(self.nb)
        self.nb.add(self.tab_prev, text="👁️ Previsualización")

        self.tab_prob = ttk.Frame(self.nb)
        self.nb.add(self.tab_prob, text="⚠️ Problemas detectados (0)")

        self._build_tab_prev()
        self._build_tab_prob()

        # Barra inferior
        frm_bot = ttk.Frame(self.win)
        frm_bot.pack(fill=tk.X, padx=16, pady=(0, 12))

        self.lbl_conteo = ttk.Label(frm_bot, text="", foreground='#64748B')
        self.lbl_conteo.pack(side=tk.LEFT)

        self.btn_importar = ttk.Button(frm_bot,
                                        text="✅ Confirmar Importación",
                                        command=self._importar,
                                        style='Create.TButton',
                                        state='disabled')
        self.btn_importar.pack(side=tk.RIGHT)

    def _build_tab_prev(self):
        frm = ttk.Frame(self.tab_prev)
        frm.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        sb_y = ttk.Scrollbar(frm); sb_y.pack(side=tk.RIGHT, fill=tk.Y)
        sb_x = ttk.Scrollbar(frm, orient=tk.HORIZONTAL)
        sb_x.pack(side=tk.BOTTOM, fill=tk.X)

        self.tree = ttk.Treeview(
            frm, columns=self.COLUMNAS, show='headings',
            yscrollcommand=sb_y.set, xscrollcommand=sb_x.set, height=12)
        sb_y.config(command=self.tree.yview)
        sb_x.config(command=self.tree.xview)

        anchos = {'codigo': 100, 'nombre': 180, 'descripcion': 180,
                  'cantidad': 70, 'precio_compra': 95, 'precio_unitario': 95,
                  'stock_minimo': 85, 'stock_maximo': 85,
                  'unidad_medida': 100, 'proveedor': 140, 'categoria': 100}
        for c in self.COLUMNAS:
            self.tree.heading(c, text=c.replace('_', ' ').title())
            self.tree.column(c, width=anchos.get(c, 100),
                             minwidth=anchos.get(c, 80))

        self.tree.tag_configure('dup_excel', background='#FEF9C3', foreground='#854D0E')
        self.tree.tag_configure('ok',        background='#F0FDF4')
        self.tree.pack(fill=tk.BOTH, expand=True)
        self.tree.pack(fill=tk.BOTH, expand=True)
        bloquear_columnas(self.tree)

    def _build_tab_prob(self):
        frm = ttk.Frame(self.tab_prob)
        frm.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        sb = ttk.Scrollbar(frm); sb.pack(side=tk.RIGHT, fill=tk.Y)
        cols = ('Fila Excel', 'Tipo', 'Detalle')
        self.tree_prob = ttk.Treeview(frm, columns=cols, show='headings',
                                       yscrollcommand=sb.set, height=14)
        sb.config(command=self.tree_prob.yview)
        for c, w in zip(cols, [90, 180, 480]):
            self.tree_prob.heading(c, text=c)
            self.tree_prob.column(c, width=w)
        self.tree_prob.tag_configure('error', background='#FEF2F2', foreground='#991B1B')
        self.tree_prob.tag_configure('warn',  background='#FEF9C3', foreground='#854D0E')
        self.tree_prob.pack(fill=tk.BOTH, expand=True)
        self.tree_prob.pack(fill=tk.BOTH, expand=True)
        bloquear_columnas(self.tree_prob)

    # ── Plantilla ─────────────────────────────────────────────────────────────

    def _descargar_plantilla(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            ruta = filedialog.asksaveasfilename(
                parent=self.win,
                defaultextension='.xlsx',
                filetypes=[("Excel", "*.xlsx")],
                initialfile="plantilla_importar_productos.xlsx")
            if not ruta: return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Productos"

            fill   = PatternFill("solid", fgColor="1f4788")
            fuente = Font(bold=True, color="FFFFFF")
            for ci, col in enumerate(self.COLUMNAS, 1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.fill = fill; cell.font = fuente
                cell.alignment = Alignment(horizontal='center')
            # Dos filas de ejemplo
            ejemplos = self._ejemplos_plantilla()
            for ri, ej in enumerate(ejemplos, 2):
                for ci, col in enumerate(self.COLUMNAS, 1):
                    ws.cell(row=ri, column=ci, value=ej.get(col, ''))

            # Nota en fila 5
            ws.cell(row=5, column=1,
                    value="* codigo: opcional pero recomendado para evitar duplicados. "
                          "nombre y precio_unitario son obligatorios. Los campos opcionales activos "
                          "se incluyen automáticamente en esta plantilla.")
            ws.cell(row=5, column=1).font = Font(italic=True, color='888888')

            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 22
            wb.save(ruta)
            messagebox.showinfo("✅", f"Plantilla guardada en:\n{ruta}", parent=self.win)
        except Exception as e:
            messagebox.showerror("❌ Error", str(e), parent=self.win)

    # ── Cargar y validar ──────────────────────────────────────────────────────

    def _cargar_archivo(self):
        ruta = filedialog.askopenfilename(
            parent=self.win,
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")])
        if not ruta: return

        try:
            import openpyxl
            wb = openpyxl.load_workbook(ruta, data_only=True)
            ws = wb.active

            headers = [
                self._mapear_cabecera(ws.cell(1, c).value)
                for c in range(1, ws.max_column + 1)
            ]

            # Validar columnas mínimas
            if 'nombre' not in headers and 'codigo' not in headers:
                messagebox.showerror(
                    "❌ Formato incorrecto",
                    "El archivo debe tener al menos las columnas:\n"
                    "  nombre  y  precio_unitario\n\n"
                    "Descarga la plantilla para ver el formato correcto.",
                    parent=self.win)
                return
            if 'precio_unitario' not in headers:
                messagebox.showerror(
                    "❌ Falta columna",
                    "No se encontró la columna 'precio_unitario'.",
                    parent=self.win)
                return

            # Leer todas las filas
            filas_raw = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(v for v in row if v is not None and str(v).strip()):
                    continue
                fila = {}
                for ci, h in enumerate(headers):
                    if h:
                        valor = row[ci] if ci < len(row) else ''
                        fila[h] = self._normalizar_valor_excel(valor)
                filas_raw.append(fila)

        except Exception as e:
            messagebox.showerror("❌ Error al leer el archivo", str(e), parent=self.win)
            return

        # ── Validar y detectar problemas ──────────────────────────────────────
        self._datos     = []
        self._problemas = []

        vistos_codigo = {}   # codigo → primera fila que lo usó
        vistos_nombre = {}   # nombre → primera fila que lo usó

        for i, fila in enumerate(filas_raw, start=2):
            nombre = str(fila.get('nombre', '') or '').strip()
            codigo = str(fila.get('codigo', '') or '').strip()

            # Validar nombre
            if not nombre and not codigo:
                self._problemas.append((i, '❌ Error', 'Fila sin nombre ni código — se omitirá'))
                continue
            if not nombre:
                nombre = f"(sin nombre — código: {codigo})"

            # Validar precio
            try:
                precio = float(fila.get('precio_unitario', 0) or 0)
                if precio < 0:
                    raise ValueError
            except (ValueError, TypeError):
                self._problemas.append((i, '❌ Error',
                    f"'{nombre}' — precio_unitario inválido: "
                    f"'{fila.get('precio_unitario', '')}'. Se omitirá."))
                continue

            # Validar cantidad
            try:
                cant = int(float(fila.get('cantidad', 0) or 0))
                if cant < 0:
                    raise ValueError
            except (ValueError, TypeError):
                self._problemas.append((i, '⚠️ Aviso',
                    f"'{nombre}' — cantidad inválida, se usará 0."))
                fila['cantidad'] = 0

            # Detectar duplicados dentro del Excel
            dup_en_excel = False
            if codigo and codigo in vistos_codigo:
                self._problemas.append((i, '⚠️ Dup. en Excel',
                    f"Código '{codigo}' ya aparece en fila {vistos_codigo[codigo]}. "
                    f"Se omitirá esta fila."))
                dup_en_excel = True
            elif codigo:
                vistos_codigo[codigo] = i

            if nombre in vistos_nombre and not dup_en_excel:
                self._problemas.append((i, '⚠️ Dup. en Excel',
                    f"Nombre '{nombre}' ya aparece en fila {vistos_nombre[nombre]}. "
                    f"Se omitirá esta fila."))
                dup_en_excel = True
            elif nombre not in vistos_nombre:
                vistos_nombre[nombre] = i

            if dup_en_excel:
                continue

            self._datos.append(fila)

        # ── Poblar previsualización ───────────────────────────────────────────
        for r in self.tree.get_children():
            self.tree.delete(r)
        for r in self.tree_prob.get_children():
            self.tree_prob.delete(r)

        nombres_con_prob = {
            str(p.get('nombre', '') or '').strip()
            for _, tip, det in self._problemas
            for p in [{}] if 'Dup' in tip
        }

        for p in self._datos[:50]:
            n   = str(p.get('nombre', '') or '').strip()
            tag = 'dup_excel' if n in nombres_con_prob else 'ok'
            self.tree.insert('', tk.END, tags=(tag,),
                             values=[p.get(c, '') for c in self.COLUMNAS])

        for fila_n, tipo, detalle in self._problemas:
            tag = 'error' if '❌' in tipo else 'warn'
            self.tree_prob.insert('', tk.END, tags=(tag,),
                                   values=(fila_n, tipo, detalle))

        # Actualizar pestaña de problemas
        n_prob = len(self._problemas)
        self.nb.tab(self.tab_prob,
                    text=f"⚠️ Problemas detectados ({n_prob})")
        if n_prob > 0:
            self.nb.select(self.tab_prob)   # abrir pestaña si hay problemas

        nombre_archivo = os.path.basename(ruta)
        self.lbl_archivo.config(text=f"✅  {nombre_archivo}", foreground='#065F46')

        total    = len(self._datos)
        omitidos = len([p for _, t, _ in self._problemas if 'Dup' in t or '❌' in t])
        txt = f"  {total} producto(s) listos para importar"
        if omitidos:
            txt += f"  |  {omitidos} fila(s) con problemas (ver pestaña ⚠️)"
        self.lbl_conteo.config(text=txt)
        self.btn_importar.config(state='normal' if total > 0 else 'disabled')

    # ── Importar ──────────────────────────────────────────────────────────────

    def _importar(self):
        if not self._datos:
            messagebox.showwarning("⚠️", "No hay datos válidos para importar.",
                                   parent=self.win); return

        n_prob = len(self._problemas)
        aviso  = ""
        if n_prob:
            aviso = (f"\n\n⚠️ Hay {n_prob} problema(s) detectado(s) en el archivo.\n"
                     f"Solo se importarán las filas válidas.")

        if not messagebox.askyesno(
                "📥 Confirmar importación",
                f"¿Importar {len(self._datos)} producto(s)?\n\n"
                f"• Productos nuevos → se insertarán\n"
                f"• Duplicados en BD (por código o nombre) → se ignorarán"
                f"{aviso}",
                parent=self.win):
            return

        # Contar antes para calcular cuántos son nuevos
        self._cats_antes = self.db.contar_categorias()
        self._provs_antes = self.db.contar_proveedores()

        insertados, actualizados, omitidos, errores = self.db.importar_productos_excel(
            self._datos, modo_duplicados='saltar')

        # Contar categorías y proveedores nuevos para el resumen
        cats_antes   = getattr(self, '_cats_antes',   0)
        provs_antes  = getattr(self, '_provs_antes',  0)
        cats_ahora = self.db.contar_categorias()
        provs_ahora = self.db.contar_proveedores()
        cats_nuevas = max(0, cats_ahora - cats_antes)
        provs_nuevos = max(0, provs_ahora - provs_antes)

        self.db.registrar_log(
            self.usuario['id'], self.usuario['username'],
            'Importar productos Excel',
            f'{insertados} nuevos, {omitidos} omitidos, {len(errores)} errores')

        resumen = f"✅ {insertados} producto(s) nuevos agregados al inventario."
        if omitidos:
            resumen += (f"\n⏭️ {omitidos} ya existían en la BD — ignorados.")
        if cats_nuevas:
            resumen += f"\n🏷️ {cats_nuevas} categoría(s) nueva(s) creadas automáticamente."
        if provs_nuevos:
            resumen += (f"\n🏭 {provs_nuevos} proveedor(es) nuevo(s) creados automáticamente."
                        f"\n   Completa sus datos en Administración → Gestionar Proveedores.")
        if errores:
            resumen += f"\n\n❌ {len(errores)} fila(s) con error de inserción:"
            for fila, msg in errores[:8]:
                resumen += f"\n  • Fila {fila}: {msg}"
            if len(errores) > 8:
                resumen += f"\n  ... y {len(errores) - 8} más"

        messagebox.showinfo("📥 Resultado", resumen, parent=self.win)
        if insertados > 0:
            self.win.destroy()
