from gui.estilos import ventana_fullscreen
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
from gui.ui_helpers import bloquear_columnas

class MovimientosWindow:
    """
    Búsqueda y filtrado de movimientos en tiempo real.
    Filtros: producto, tipo, categoría, rango de fechas.
    Columnas ordenables haciendo clic en el encabezado.
    """

    def __init__(self, master, db, C, usuario=None):
        self.db      = db
        self.C       = C
        self.usuario = usuario
        self.can_export = bool(
            usuario and (
                db.tiene_permiso(usuario, 'exportar_movimientos') or
                db.tiene_permiso(usuario, 'exportar_todo')
            )
        )

        self.win = ventana_fullscreen(master, "🔎 Buscar Movimientos", C)
        self._build()
        self._buscar()

    def _build(self):
        C = self.C

        # Encabezado
        hdr = tk.Frame(self.win, bg=C['header_bg'], height=50)
        hdr.pack(fill=tk.X); hdr.pack_propagate(False)
        tk.Label(hdr, text="🔎  Búsqueda de Movimientos",
                 font=("Segoe UI", 13, "bold"),
                 bg=C['header_bg'], fg='white').pack(side=tk.LEFT, padx=16)
        body = tk.Frame(self.win, bg=C['bg'])
        body.pack(fill=tk.BOTH, expand=True, padx=16, pady=12)

        # ── Filtros ───────────────────────────────────────────────────────────
        frm_f = tk.LabelFrame(body, text="  Filtros  ",
                               bg=C['bg'], fg=C['text'],
                               font=('Segoe UI', 9, 'bold'))
        frm_f.pack(fill=tk.X, pady=(0, 10))

        row1 = tk.Frame(frm_f, bg=C['bg']); row1.pack(fill=tk.X, padx=10, pady=8)

        def lbl(parent, txt):
            tk.Label(parent, text=txt, font=('Segoe UI', 9),
                     bg=C['bg'], fg=C['muted']).pack(side=tk.LEFT, padx=(0, 4))

        lbl(row1, "Producto / Nota:")
        self.e_buscar = ttk.Entry(row1, width=24)
        self.e_buscar.pack(side=tk.LEFT, padx=(0, 16))
        self.e_buscar.bind('<KeyRelease>', lambda e: self._buscar())

        lbl(row1, "Tipo:")
        self.cb_tipo = ttk.Combobox(row1, values=['Todos', 'Entrada', 'Salida', 'Ajuste'],
                                     state='readonly', width=10)
        self.cb_tipo.set('Todos'); self.cb_tipo.pack(side=tk.LEFT, padx=(0, 16))
        self.cb_tipo.bind('<<ComboboxSelected>>', lambda e: self._buscar())

        lbl(row1, "Categoría:")
        cats = ['Todas'] + self.db.obtener_nombres_categorias()
        self.cb_cat = ttk.Combobox(row1, values=cats, state='readonly', width=16)
        self.cb_cat.set('Todas'); self.cb_cat.pack(side=tk.LEFT, padx=(0, 16))
        self.cb_cat.bind('<<ComboboxSelected>>', lambda e: self._buscar())

        row2 = tk.Frame(frm_f, bg=C['bg']); row2.pack(fill=tk.X, padx=10, pady=(0, 8))

        lbl(row2, "Desde:")
        self.e_desde = ttk.Entry(row2, width=13)
        self.e_desde.insert(0, (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        self.e_desde.pack(side=tk.LEFT, padx=(0, 12))
        self.e_desde.bind('<FocusOut>', lambda e: self._buscar())

        lbl(row2, "Hasta:")
        self.e_hasta = ttk.Entry(row2, width=13)
        self.e_hasta.insert(0, datetime.now().strftime('%Y-%m-%d'))
        self.e_hasta.pack(side=tk.LEFT, padx=(0, 12))
        self.e_hasta.bind('<FocusOut>', lambda e: self._buscar())

        ttk.Button(row2, text="✖ Limpiar filtros",
                   command=self._limpiar,
                   style='Neutral.TButton').pack(side=tk.LEFT, padx=(8, 0))

        if self.can_export:
            ttk.Button(row2, text="📥 Exportar Excel",
                       command=self._exportar_excel,
                       style='Neutral.TButton').pack(side=tk.LEFT, padx=(8, 0))

            ttk.Button(row2, text="📄 Exportar PDF",
                       command=self._exportar_pdf,
                       style='Neutral.TButton').pack(side=tk.LEFT, padx=(4, 0))

        self.lbl_total = tk.Label(row2, text="", font=('Segoe UI', 9),
                                   bg=C['bg'], fg=C['muted'])
        self.lbl_total.pack(side=tk.RIGHT, padx=8)

        # ── Tabla ─────────────────────────────────────────────────────────────
        frm_t = tk.Frame(body, bg=C['bg'])
        frm_t.pack(fill=tk.BOTH, expand=True)

        sb_y = ttk.Scrollbar(frm_t); sb_y.pack(side=tk.RIGHT, fill=tk.Y)
        sb_x = ttk.Scrollbar(frm_t, orient=tk.HORIZONTAL)
        sb_x.pack(side=tk.BOTTOM, fill=tk.X)

        cols = ('ID', 'Producto', 'Categoría', 'Tipo', 'Cantidad',
                'Fecha', 'Usuario', 'Nota')
        self.tree = ttk.Treeview(frm_t, columns=cols, show='headings',
                                  yscrollcommand=sb_y.set,
                                  xscrollcommand=sb_x.set)
        sb_y.config(command=self.tree.yview)
        sb_x.config(command=self.tree.xview)

        anchos = {'ID': 50, 'Producto': 200, 'Categoría': 120, 'Tipo': 90,
                  'Cantidad': 80, 'Fecha': 150, 'Usuario': 120, 'Nota': 220}
        for c in cols:
            self.tree.heading(c, text=c,
                               command=lambda col=c: self._ordenar(col))
            self.tree.column(c, width=anchos[c])

        self.tree.tag_configure('entrada',
                                background='#F0FDF4', foreground='#065F46')
        self.tree.tag_configure('salida',
                                background='#FEF2F2', foreground='#991B1B')
        self.tree.pack(fill=tk.BOTH, expand=True)
        self.tree.pack(fill=tk.BOTH, expand=True)
        bloquear_columnas(self.tree)

        self._orden_col = None
        self._orden_asc = True

    # ── Lógica ────────────────────────────────────────────────────────────────

    def _buscar(self):
        termino   = self.e_buscar.get().strip()
        tipo      = self.cb_tipo.get() if self.cb_tipo.get() != 'Todos'  else ''
        categoria = self.cb_cat.get()  if self.cb_cat.get()  != 'Todas'  else ''
        desde     = self.e_desde.get().strip()
        hasta     = self.e_hasta.get().strip()
        movs      = self.db.buscar_movimientos(termino, tipo, categoria, desde, hasta)
        self._poblar(movs)

    def _poblar(self, movs):
        self._movs_actuales = movs
        for row in self.tree.get_children(): self.tree.delete(row)
        for m in movs:
            tipo  = str(m.get('tipo_movimiento', '')).lower()
            tag   = 'entrada' if 'entrada' in tipo else 'salida'
            fecha = m['fecha'].strftime('%d/%m/%Y %H:%M') if m.get('fecha') else ''
            self.tree.insert('', tk.END, tags=(tag,), values=(
                m.get('id', ''),
                m.get('nombre_producto', ''),
                m.get('categoria_producto', ''),
                m.get('tipo_movimiento', ''),
                m.get('cantidad', 0),
                fecha,
                m.get('usuario_nombre') or 'Sistema',
                m.get('descripcion', '') or ''))
        self.lbl_total.config(text=f"  {len(movs)} resultado(s)")

    def _ordenar(self, col):
        items   = [(self.tree.set(k, col), k) for k in self.tree.get_children('')]
        reverse = (self._orden_col == col and self._orden_asc)
        try:
            items.sort(
                key=lambda x: float(x[0]) if x[0].replace('.','',1).isdigit()
                              else x[0].lower(),
                reverse=reverse)
        except Exception:
            items.sort(reverse=reverse)
        for i, (_, k) in enumerate(items): self.tree.move(k, '', i)
        self._orden_col = col
        self._orden_asc = not reverse

    def _limpiar(self):
        self.e_buscar.delete(0, tk.END)
        self.cb_tipo.set('Todos')
        self.cb_cat.set('Todas')
        self.e_desde.delete(0, tk.END)
        self.e_desde.insert(0, (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        self.e_hasta.delete(0, tk.END)
        self.e_hasta.insert(0, datetime.now().strftime('%Y-%m-%d'))
        self._buscar()

    def _exportar_excel(self):
        if not self.can_export:
            messagebox.showwarning("Permiso requerido", "No tienes permiso para exportar movimientos.", parent=self.win)
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from tkinter import filedialog
            from datetime import datetime

            movs = getattr(self, '_movs_actuales', [])
            if not movs:
                from tkinter import messagebox
                messagebox.showwarning("⚠️", "No hay movimientos para exportar.", parent=self.win)
                return

            ruta = filedialog.asksaveasfilename(
                parent=self.win, defaultextension='.xlsx',
                filetypes=[('Excel', '*.xlsx')],
                initialfile=f"movimientos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            if not ruta: return

            wb = openpyxl.Workbook()
            ws = wb.active; ws.title = "Movimientos"
            hdrs = ['ID', 'Producto', 'Categoría', 'Tipo', 'Cantidad', 'Fecha', 'Usuario', 'Nota']
            hfill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            for i, h in enumerate(hdrs, 1):
                cell = ws.cell(row=1, column=i, value=h)
                cell.fill = hfill
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center')

            for ri, m in enumerate(movs, 2):
                vals = [m.get('id'), m.get('nombre_producto',''),
                        m.get('categoria_producto',''), m.get('tipo_movimiento',''),
                        m.get('cantidad',0),
                        str(m['fecha'])[:19] if m.get('fecha') else '',
                        m.get('usuario_nombre','') or 'Sistema',
                        m.get('descripcion','') or '']
                for ci, v in enumerate(vals, 1):
                    ws.cell(row=ri, column=ci, value=v)

            for col in ws.columns:
                ml = max((len(str(c.value or '')) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(50, ml+3)

            wb.save(ruta)
            from tkinter import messagebox
            messagebox.showinfo("✅", f"Exportado:\n{ruta}", parent=self.win)
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("❌", str(e), parent=self.win)

    def _exportar_pdf(self):
        if not self.can_export:
            messagebox.showwarning("Permiso requerido", "No tienes permiso para exportar movimientos.", parent=self.win)
            return
        try:
            from tkinter import filedialog, messagebox
            from datetime import datetime
            ruta = filedialog.asksaveasfilename(
                parent=self.win, defaultextension='.pdf',
                filetypes=[('PDF', '*.pdf')],
                initialfile=f"movimientos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
            if not ruta: return

            movs = getattr(self, '_movs_actuales', [])
            if not movs:
                messagebox.showwarning("⚠️", "No hay movimientos.", parent=self.win); return

            from reportes_gen.reports import ReportGenerator
            prods = {p['id']: p for p in self.db.obtener_productos()}
            ok, msg = ReportGenerator().generar_reporte_movimientos(movs, prods,
                titulo_base="Movimientos Filtrados", ruta_destino=ruta)
            if ok:
                messagebox.showinfo("✅", f"PDF guardado:\n{ruta}", parent=self.win)
            else:
                messagebox.showerror("❌", msg, parent=self.win)
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("❌", str(e), parent=self.win)
