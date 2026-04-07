import tkinter as tk
from tkinter import ttk
from gui.ui_helpers import bloquear_columnas, configurar_ventana

class LogWindow:
    """Ventana para ver el historial/auditoría del sistema."""

    def __init__(self, master, db):
        self.db = db

        self.win = tk.Toplevel(master)
        self.win.title("🧾 Historial de Actividad del Sistema")
        self.win.configure(bg='#F1F5F9')
        self.win.grab_set()
        configurar_ventana(self.win, width=1120, height=700, min_width=980, min_height=560)

        self._build()
        self._cargar()

    def _build(self):
        ttk.Label(self.win, text="🧾 Auditoría — Registro de Actividad",
                  font=('Segoe UI', 13, 'bold'), background='#F1F5F9',
                  foreground='#2563EB').pack(pady=(12, 4), padx=14)
        ttk.Label(self.win,
                  text="Muestra las últimas 500 acciones registradas en el sistema.",
                  font=('Segoe UI', 9), background='#F1F5F9',
                  foreground='#64748B').pack(padx=14, pady=(0, 8))

        # Barra de filtro
        frm_f = ttk.Frame(self.win)
        frm_f.pack(fill=tk.X, padx=12, pady=(0, 4))

        # Fila 1: búsqueda + fechas
        row1 = ttk.Frame(frm_f); row1.pack(fill=tk.X, pady=(0,4))
        ttk.Label(row1, text="Buscar:").pack(side=tk.LEFT, padx=(0, 4))
        self.e_filtro = ttk.Entry(row1, width=18)
        self.e_filtro.pack(side=tk.LEFT, padx=(0, 10))
        self.e_filtro.bind('<KeyRelease>', self._filtrar)

        ttk.Label(row1, text="Desde:").pack(side=tk.LEFT, padx=(0, 4))
        self.e_desde = ttk.Entry(row1, width=12)
        self.e_desde.pack(side=tk.LEFT, padx=(0, 8))
        self.e_desde.bind('<FocusOut>', self._filtrar)

        ttk.Label(row1, text="Hasta:").pack(side=tk.LEFT, padx=(0, 4))
        self.e_hasta = ttk.Entry(row1, width=12)
        self.e_hasta.pack(side=tk.LEFT, padx=(0, 10))
        self.e_hasta.bind('<FocusOut>', self._filtrar)

        ttk.Button(row1, text="🔄 Recargar", command=self._cargar).pack(side=tk.LEFT, padx=4)

        # Fila 2: exportación
        row2 = ttk.Frame(frm_f); row2.pack(fill=tk.X, pady=(0,4))
        ttk.Button(row2, text="📥 Exportar a Excel",
                   command=self._exportar_excel,
                   style='Neutral.TButton').pack(side=tk.LEFT, padx=(0, 6))
        self.lbl_count = ttk.Label(row2, text="", foreground='#64748B')
        self.lbl_count.pack(side=tk.LEFT)

        # Tabla
        frm = ttk.Frame(self.win)
        frm.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 12))

        sb_y = ttk.Scrollbar(frm); sb_y.pack(side=tk.RIGHT, fill=tk.Y)
        sb_x = ttk.Scrollbar(frm, orient=tk.HORIZONTAL); sb_x.pack(side=tk.BOTTOM, fill=tk.X)

        cols = ('ID', 'Usuario', 'Acción', 'Detalle', 'Producto Afectado', 'Fecha')
        self.tree = ttk.Treeview(frm, columns=cols, show='headings',
                                  yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
        sb_y.config(command=self.tree.yview)
        sb_x.config(command=self.tree.xview)

        anchos = {'ID': 45, 'Usuario': 110, 'Acción': 180,
                  'Detalle': 200, 'Producto Afectado': 160, 'Fecha': 150}
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=anchos[c],
                              anchor=tk.CENTER if c == 'ID' else tk.W)

        # Color por acción
        self.tree.tag_configure('eliminar',   background='#FEE2E2', foreground='#991B1B')
        self.tree.tag_configure('crear',      background='#D1FAE5', foreground='#065F46')
        self.tree.tag_configure('exportar',   background='#EDE9FE', foreground='#5B21B6')
        self.tree.tag_configure('movimiento', background='#DBEAFE', foreground='#1E3A8A')
        self.tree.tag_configure('login',      background='#FEF9C3', foreground='#854D0E')
        self.tree.tag_configure('default',    background='#F8FAFC')

        self.tree.pack(fill=tk.BOTH, expand=True)
        self.tree.pack(fill=tk.BOTH, expand=True)
        bloquear_columnas(self.tree)

        self._todos = []

    def _cargar(self):
        self._todos = self.db.obtener_log(500)
        self._renderizar(self._todos)

    def _filtrar(self, event=None):
        t      = self.e_filtro.get().strip().lower()
        desde  = self.e_desde.get().strip() if hasattr(self, 'e_desde') else ''
        hasta  = self.e_hasta.get().strip() if hasattr(self, 'e_hasta') else ''
        result = self._todos

        if t:
            result = [r for r in result
                      if t in str(r.get('username', '')).lower()
                      or t in str(r.get('accion', '')).lower()
                      or t in str(r.get('detalle', '')).lower()]

        if desde:
            try:
                from datetime import datetime
                d = datetime.strptime(desde, '%Y-%m-%d')
                result = [r for r in result if r.get('fecha') and r['fecha'].date() >= d.date()]
            except Exception:
                pass

        if hasta:
            try:
                from datetime import datetime
                h = datetime.strptime(hasta, '%Y-%m-%d')
                result = [r for r in result if r.get('fecha') and r['fecha'].date() <= h.date()]
            except Exception:
                pass

        self._renderizar(result)

    def _renderizar(self, registros):
        for i in self.tree.get_children():
            self.tree.delete(i)
        try:
            if hasattr(self, 'lbl_count'):
                self.lbl_count.config(text=f"{len(registros)} registro(s)")
        except Exception:
            pass
        for r in registros:
            accion = str(r.get('accion', '')).lower()
            if 'eliminar' in accion or 'delete' in accion:
                tag = 'eliminar'
            elif 'crear' in accion or 'nuevo' in accion:
                tag = 'crear'
            elif 'exportar' in accion or 'backup' in accion:
                tag = 'exportar'
            elif 'movimiento' in accion or 'entrada' in accion or 'salida' in accion:
                tag = 'movimiento'
            elif 'login' in accion or 'sesión' in accion or 'password' in accion:
                tag = 'login'
            else:
                tag = 'default'

            self.tree.insert('', tk.END, tags=(tag,), values=(
                r['id'],
                r.get('username', 'Sistema'),
                r.get('accion', ''),
                r.get('detalle', '') or '',
                r.get('producto_afectado', '') or '',
                str(r['fecha'])[:19] if r.get('fecha') else '',
            ))

    def _exportar_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from tkinter import filedialog
            from datetime import datetime

            ruta = filedialog.asksaveasfilename(
                parent=self.win,
                defaultextension='.xlsx',
                filetypes=[('Excel', '*.xlsx')],
                initialfile=f"auditoria_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            if not ruta:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Auditoría"

            # Determinar registros visibles (con filtros aplicados)
            t     = self.e_filtro.get().strip().lower() if hasattr(self, 'e_filtro') else ''
            desde = self.e_desde.get().strip() if hasattr(self, 'e_desde') else ''
            hasta = self.e_hasta.get().strip() if hasattr(self, 'e_hasta') else ''
            data  = self._todos

            if t:
                data = [r for r in data if t in str(r.get('username','')).lower()
                        or t in str(r.get('accion','')).lower()]
            if desde:
                try:
                    from datetime import datetime as dt
                    d = dt.strptime(desde, '%Y-%m-%d')
                    data = [r for r in data if r.get('fecha') and r['fecha'].date() >= d.date()]
                except Exception: pass
            if hasta:
                try:
                    from datetime import datetime as dt
                    h = dt.strptime(hasta, '%Y-%m-%d')
                    data = [r for r in data if r.get('fecha') and r['fecha'].date() <= h.date()]
                except Exception: pass

            hdrs = ['ID', 'Usuario', 'Acción', 'Detalle', 'Producto Afectado', 'Fecha']
            hfill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            hfont = Font(bold=True, color="FFFFFF")
            for i, h in enumerate(hdrs, 1):
                cell = ws.cell(row=1, column=i, value=h)
                cell.fill = hfill; cell.font = hfont
                cell.alignment = Alignment(horizontal='center')

            for row_i, r in enumerate(data, 2):
                ws.cell(row=row_i, column=1, value=r.get('id'))
                ws.cell(row=row_i, column=2, value=r.get('username', ''))
                ws.cell(row=row_i, column=3, value=r.get('accion', ''))
                ws.cell(row=row_i, column=4, value=r.get('detalle', '') or '')
                ws.cell(row=row_i, column=5, value=r.get('producto_afectado', '') or '')
                ws.cell(row=row_i, column=6, value=str(r['fecha'])[:19] if r.get('fecha') else '')

            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(50, max_len + 3)

            wb.save(ruta)
            from tkinter import messagebox
            messagebox.showinfo("✅", f"Exportado:\n{ruta}", parent=self.win)

        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("❌ Error", str(e), parent=self.win)
