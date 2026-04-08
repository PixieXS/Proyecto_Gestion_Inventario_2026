import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import REPORTS_PATH


def _slug(texto):
    limpio = "".join(ch.lower() if ch.isalnum() else "_" for ch in str(texto or "reporte"))
    while "__" in limpio:
        limpio = limpio.replace("__", "_")
    return limpio.strip("_") or "reporte"


def _safe_sheet_title(texto):
    inval = {":", "\\", "/", "?", "*", "[", "]"}
    limpio = "".join("_" if ch in inval else ch for ch in str(texto or "Hoja"))
    return limpio[:31] or "Hoja"


class ExcelExporter:
    def __init__(self, db=None):
        self.db = db
        os.makedirs(REPORTS_PATH, exist_ok=True)

    @staticmethod
    def _base_styles(color_base="1F4788"):
        return {
            "fill_header": PatternFill("solid", fgColor=color_base),
            "fill_soft": PatternFill("solid", fgColor="F8FAFC"),
            "fill_light": PatternFill("solid", fgColor="EAF2FF"),
            "fill_ok": PatternFill("solid", fgColor="ECFDF5"),
            "fill_alert": PatternFill("solid", fgColor="FEF2F2"),
            "font_header": Font(bold=True, color="FFFFFF", size=11),
            "font_title": Font(bold=True, size=16, color="1E293B"),
            "font_section": Font(bold=True, size=11, color="1E293B"),
            "font_label": Font(bold=True, size=10, color="1E293B"),
            "font_muted": Font(italic=True, color="64748B"),
            "border": Border(
                left=Side(style="thin", color="CBD5E1"),
                right=Side(style="thin", color="CBD5E1"),
                top=Side(style="thin", color="CBD5E1"),
                bottom=Side(style="thin", color="CBD5E1"),
            ),
            "align_wrap": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "align_center": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "align_right": Alignment(horizontal="right", vertical="top", wrap_text=True),
        }

    @staticmethod
    def _texto_filtros(filtros):
        etiquetas = {
            "fecha_inicio": "Fecha inicio",
            "fecha_fin": "Fecha fin",
            "categoria": "Categoria",
            "id_proveedor": "Proveedor",
            "id_producto": "Producto",
            "tipo_movimiento": "Tipo movimiento",
            "id_usuario": "Usuario",
            "estado_producto": "Estado producto",
        }
        filas = []
        for clave, valor in (filtros or {}).items():
            if valor in (None, "", "todos"):
                continue
            filas.append((etiquetas.get(clave, clave), valor))
        return filas

    @staticmethod
    def _formato_por_fila(reporte, columna, fila):
        formatos = dict(reporte.get("formatos") or {})
        fmt = formatos.get(columna)
        if fmt == "auto" and reporte.get("tipo") == "comparativo_fechas":
            metrica = str(fila.get("Metrica") or "").lower()
            return "currency" if "valor" in metrica else "number"
        if fmt == "metric":
            valor = fila.get(columna)
            return "currency" if isinstance(valor, float) else "number"
        return fmt

    @staticmethod
    def _aplicar_formato_excel(cell, fmt):
        if fmt == "currency":
            cell.number_format = "$#,##0.00"
        elif fmt == "percent":
            cell.number_format = '0.0"%"'
        elif fmt == "date":
            cell.number_format = "dd/mm/yyyy"
        elif fmt == "datetime":
            cell.number_format = "dd/mm/yyyy hh:mm"
        elif fmt == "number":
            cell.number_format = "#,##0"

    @staticmethod
    def _auto_fit_columns(ws, min_width=10, max_width=42):
        for idx, column_cells in enumerate(ws.iter_cols(1, ws.max_column), start=1):
            max_len = 0
            for cell in column_cells:
                valor = cell.value
                if valor is None:
                    continue
                texto = valor.strftime("%d/%m/%Y %H:%M") if hasattr(valor, "strftime") else str(valor)
                line_len = max((len(line) for line in texto.splitlines()), default=0)
                max_len = max(max_len, line_len)
            ws.column_dimensions[get_column_letter(idx)].width = max(min_width, min(max_width, max_len + 3))

    @staticmethod
    def _adjust_row_heights(ws, start_row=1, end_row=None, min_height=18):
        end_row = end_row or ws.max_row
        for row in range(start_row, end_row + 1):
            max_lines = 1
            for cell in ws[row]:
                if cell.value in (None, ""):
                    continue
                texto = cell.value.strftime("%d/%m/%Y %H:%M") if hasattr(cell.value, "strftime") else str(cell.value)
                max_lines = max(max_lines, len(texto.splitlines()))
                if len(texto) > 45:
                    max_lines = max(max_lines, int(len(texto) / 35) + 1)
            ws.row_dimensions[row].height = max(min_height, min(60, max_lines * 16))

    def _write_kv_table(self, ws, start_row, title, rows, styles):
        ws.cell(row=start_row, column=1, value=title).font = styles["font_section"]
        start_row += 1
        ws.cell(row=start_row, column=1, value="Campo")
        ws.cell(row=start_row, column=2, value="Valor")
        for col in (1, 2):
            cell = ws.cell(row=start_row, column=col)
            cell.fill = styles["fill_header"]
            cell.font = styles["font_header"]
            cell.alignment = styles["align_center"]
            cell.border = styles["border"]
        start_row += 1
        for idx, (label, value) in enumerate(rows):
            for col, item in enumerate((label, value), start=1):
                cell = ws.cell(row=start_row, column=col, value=item)
                cell.border = styles["border"]
                cell.alignment = styles["align_wrap"]
                cell.fill = styles["fill_soft"] if idx % 2 else PatternFill(fill_type=None)
            start_row += 1
        return start_row

    def exportar_reporte_ejecutivo(self, reporte, filtros=None, usuario=None, empresa_info=None):
        try:
            wb = openpyxl.Workbook()
            ws_res = wb.active
            ws_res.title = "Resumen"
            ws_det = wb.create_sheet("Detalle")
            styles = self._base_styles()

            empresa = (empresa_info or {}).get("nombre") or "Inventoryx"
            ws_res["A1"] = empresa
            ws_res["A1"].font = styles["font_title"]
            ws_res["A2"] = reporte.get("titulo") or "Reporte"
            ws_res["A2"].font = Font(bold=True, size=13, color="1E293B")
            ws_res["A3"] = reporte.get("subtitulo") or ""
            ws_res["A3"].font = styles["font_muted"]
            ws_res["A5"] = "Generado"
            ws_res["B5"] = datetime.now().strftime("%d/%m/%Y %H:%M")
            ws_res["A6"] = "Usuario"
            ws_res["B6"] = (usuario or {}).get("username") or "Sistema"
            ws_res["A5"].font = styles["font_label"]
            ws_res["A6"].font = styles["font_label"]

            row = 8
            row = self._write_kv_table(ws_res, row, "Filtros aplicados", self._texto_filtros(filtros) or [("Filtros", "Sin filtros adicionales")], styles)
            row += 1
            resumen_rows = list((reporte.get("resumen") or {}).items()) or [("Resumen", "Sin indicadores")]
            row = self._write_kv_table(ws_res, row, "Resumen ejecutivo", resumen_rows, styles)
            ws_res.column_dimensions["A"].width = 30
            ws_res.column_dimensions["B"].width = 42
            self._adjust_row_heights(ws_res, 1, ws_res.max_row)

            titulo_detalle = reporte.get("titulo") or "Detalle"
            ws_det.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(reporte.get("columnas") or [])))
            ws_det["A1"] = titulo_detalle
            ws_det["A1"].font = styles["font_title"]
            ws_det["A2"] = reporte.get("subtitulo") or ""
            ws_det["A2"].font = styles["font_muted"]

            header_row = 4
            columnas = list(reporte.get("columnas") or [])
            for col_idx, columna in enumerate(columnas, start=1):
                cell = ws_det.cell(row=header_row, column=col_idx, value=columna)
                cell.fill = styles["fill_header"]
                cell.font = styles["font_header"]
                cell.alignment = styles["align_center"]
                cell.border = styles["border"]

            for row_idx, fila_data in enumerate(reporte.get("filas") or [], start=header_row + 1):
                stock = fila_data.get("Stock")
                stock_min = fila_data.get("Stock Minimo")
                balance = fila_data.get("Balance")
                fill = None
                if stock is not None and stock_min is not None and float(stock) <= float(stock_min):
                    fill = styles["fill_alert"]
                elif balance is not None and float(balance) >= 0:
                    fill = styles["fill_ok"]

                for col_idx, columna in enumerate(columnas, start=1):
                    valor = fila_data.get(columna)
                    cell = ws_det.cell(row=row_idx, column=col_idx, value=valor)
                    cell.border = styles["border"]
                    fmt = self._formato_por_fila(reporte, columna, fila_data)
                    self._aplicar_formato_excel(cell, fmt)
                    if fmt in {"currency", "number", "percent"}:
                        cell.alignment = styles["align_right"]
                    else:
                        cell.alignment = styles["align_wrap"]
                    if fill:
                        cell.fill = fill

            if columnas:
                end_col = get_column_letter(len(columnas))
                ws_det.auto_filter.ref = f"A{header_row}:{end_col}{max(header_row, ws_det.max_row)}"
                ws_det.freeze_panes = f"A{header_row + 1}"
            self._auto_fit_columns(ws_det, min_width=10, max_width=40)
            self._adjust_row_heights(ws_det, header_row, ws_det.max_row)

            nombre = reporte.get("nombre_archivo") or _slug(reporte.get("titulo") or "reporte")
            filename = f"{REPORTS_PATH}{nombre}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            return True, filename
        except Exception as err:
            return False, f"Error al exportar reporte ejecutivo: {str(err)}"

    def exportar_inventario(self, productos):
        reporte = {
            "titulo": "Inventario",
            "columnas": ["ID", "Codigo", "Nombre", "Categoria", "Cantidad", "Precio Unitario", "Valor Total", "Proveedor"],
            "filas": [
                {
                    "ID": producto.get("id"),
                    "Codigo": producto.get("codigo") or "",
                    "Nombre": producto.get("nombre", "N/A"),
                    "Categoria": producto.get("categoria", "General"),
                    "Cantidad": producto.get("cantidad", 0),
                    "Precio Unitario": float(producto.get("precio_unitario", 0)),
                    "Valor Total": float(producto.get("cantidad", 0)) * float(producto.get("precio_unitario", 0)),
                    "Proveedor": producto.get("proveedor", "N/A"),
                }
                for producto in productos
            ],
            "formatos": {"Cantidad": "number", "Precio Unitario": "currency", "Valor Total": "currency"},
            "resumen": {"Productos": len(productos)},
            "nombre_archivo": "inventario",
        }
        return self.exportar_reporte_ejecutivo(reporte)

    def exportar_movimientos(self, movimientos, productos_dict):
        filas = []
        for mov in movimientos:
            producto_nombre = (productos_dict.get(mov.get("id_producto"), {}).get("nombre", "N/A") if productos_dict else "N/A")
            filas.append(
                {
                    "ID Movimiento": mov.get("id"),
                    "Producto": producto_nombre,
                    "Tipo": mov.get("tipo_movimiento", "N/A"),
                    "Cantidad": mov.get("cantidad", 0),
                    "Fecha": mov.get("fecha", ""),
                    "Descripcion": mov.get("descripcion", ""),
                }
            )
        reporte = {
            "titulo": "Movimientos",
            "columnas": ["ID Movimiento", "Producto", "Tipo", "Cantidad", "Fecha", "Descripcion"],
            "filas": filas,
            "formatos": {"Cantidad": "number", "Fecha": "datetime"},
            "resumen": {"Movimientos": len(filas)},
            "nombre_archivo": "movimientos",
        }
        return self.exportar_reporte_ejecutivo(reporte)

    def exportar_completo(self, productos, movimientos, productos_dict):
        try:
            wb = openpyxl.Workbook()
            styles = self._base_styles()
            ws_res = wb.active
            ws_res.title = "Resumen"
            ws_inv = wb.create_sheet("Inventario")
            ws_mov = wb.create_sheet("Movimientos")

            ws_res["A1"] = "Resumen general"
            ws_res["A1"].font = styles["font_title"]
            ws_res["A3"] = "Fecha"
            ws_res["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M")
            ws_res["A4"] = "Productos"
            ws_res["B4"] = len(productos)
            ws_res["A5"] = "Movimientos"
            ws_res["B5"] = len(movimientos)
            ws_res["A6"] = "Stock total"
            ws_res["B6"] = sum(int(p.get("cantidad", 0)) for p in productos)
            ws_res["A7"] = "Valor total"
            ws_res["B7"] = sum(float(p.get("cantidad", 0)) * float(p.get("precio_unitario", 0)) for p in productos)
            ws_res["B7"].number_format = "$#,##0.00"
            ws_res.column_dimensions["A"].width = 24
            ws_res.column_dimensions["B"].width = 18

            inv_headers = ["ID", "Codigo", "Nombre", "Categoria", "Cantidad", "Precio Unitario", "Valor Total", "Proveedor"]
            for col_idx, header in enumerate(inv_headers, start=1):
                cell = ws_inv.cell(row=1, column=col_idx, value=header)
                cell.fill = styles["fill_header"]
                cell.font = styles["font_header"]
                cell.alignment = styles["align_center"]
                cell.border = styles["border"]
            for row_idx, producto in enumerate(productos, start=2):
                valores = [
                    producto.get("id"),
                    producto.get("codigo") or "",
                    producto.get("nombre", ""),
                    producto.get("categoria", "General"),
                    producto.get("cantidad", 0),
                    float(producto.get("precio_unitario", 0)),
                    float(producto.get("cantidad", 0)) * float(producto.get("precio_unitario", 0)),
                    producto.get("proveedor", "N/A"),
                ]
                for col_idx, valor in enumerate(valores, start=1):
                    cell = ws_inv.cell(row=row_idx, column=col_idx, value=valor)
                    cell.border = styles["border"]
                    cell.alignment = styles["align_right"] if col_idx in {5, 6, 7} else styles["align_wrap"]
                    if col_idx in {5}:
                        cell.number_format = "#,##0"
                    if col_idx in {6, 7}:
                        cell.number_format = "$#,##0.00"

            mov_headers = ["ID Movimiento", "Producto", "Tipo", "Cantidad", "Fecha", "Descripcion"]
            for col_idx, header in enumerate(mov_headers, start=1):
                cell = ws_mov.cell(row=1, column=col_idx, value=header)
                cell.fill = styles["fill_header"]
                cell.font = styles["font_header"]
                cell.alignment = styles["align_center"]
                cell.border = styles["border"]
            for row_idx, mov in enumerate(movimientos, start=2):
                producto_nombre = (productos_dict.get(mov.get("id_producto"), {}).get("nombre", "N/A") if productos_dict else "N/A")
                valores = [mov.get("id"), producto_nombre, mov.get("tipo_movimiento", ""), mov.get("cantidad", 0), mov.get("fecha", ""), mov.get("descripcion", "")]
                for col_idx, valor in enumerate(valores, start=1):
                    cell = ws_mov.cell(row=row_idx, column=col_idx, value=valor)
                    cell.border = styles["border"]
                    cell.alignment = styles["align_right"] if col_idx == 4 else styles["align_wrap"]
                    if col_idx == 4:
                        cell.number_format = "#,##0"
                    if col_idx == 5 and hasattr(valor, "strftime"):
                        cell.number_format = "dd/mm/yyyy hh:mm"

            ws_inv.freeze_panes = "A2"
            ws_mov.freeze_panes = "A2"
            self._auto_fit_columns(ws_inv, min_width=10, max_width=38)
            self._auto_fit_columns(ws_mov, min_width=10, max_width=38)
            self._adjust_row_heights(ws_inv, 1, ws_inv.max_row)
            self._adjust_row_heights(ws_mov, 1, ws_mov.max_row)

            filename = f"{REPORTS_PATH}inventario_completo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            return True, filename
        except Exception as err:
            return False, f"Error al exportar datos completos: {str(err)}"

    def exportar_inventario_con_resumen(self, productos, empresa_nombre="Inventoryx"):
        reporte = {
            "titulo": f"Inventario - {empresa_nombre}",
            "columnas": ["Producto", "Categoria", "Cantidad", "Precio Unitario", "Valor Total"],
            "filas": [
                {
                    "Producto": producto.get("nombre", ""),
                    "Categoria": producto.get("categoria", "General"),
                    "Cantidad": producto.get("cantidad", 0),
                    "Precio Unitario": float(producto.get("precio_unitario", 0)),
                    "Valor Total": float(producto.get("cantidad", 0)) * float(producto.get("precio_unitario", 0)),
                }
                for producto in productos
            ],
            "formatos": {"Cantidad": "number", "Precio Unitario": "currency", "Valor Total": "currency"},
            "resumen": {
                "Productos": len(productos),
                "Stock total": sum(int(p.get("cantidad", 0)) for p in productos),
                "Valor total": sum(float(p.get("cantidad", 0)) * float(p.get("precio_unitario", 0)) for p in productos),
            },
            "nombre_archivo": "inventario_resumen",
        }
        return self.exportar_reporte_ejecutivo(reporte, empresa_info={"nombre": empresa_nombre})
